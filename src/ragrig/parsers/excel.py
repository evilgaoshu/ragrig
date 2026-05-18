from __future__ import annotations

from hashlib import sha256
from pathlib import Path

from ragrig.parsers.base import ParseResult


class ExcelParserError(ValueError):
    """Raised when an Excel file cannot be parsed."""


class ExcelParser:
    parser_name = "excel"
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def parse(self, path: Path) -> ParseResult:
        try:
            import openpyxl
        except ImportError as exc:
            raise ExcelParserError(
                "openpyxl is required for Excel parsing: pip install openpyxl"
            ) from exc

        raw_bytes = path.read_bytes()
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            raise ExcelParserError(f"Excel file could not be read: {path.name}") from exc

        sheet_texts: list[str] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First row → headers; skip if all None
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            if not any(headers):
                continue

            section_lines = [f"[Sheet: {sheet_name}]"]
            for row in rows[1:]:
                if all(v is None for v in row):
                    continue
                pairs = []
                for header, value in zip(headers, row, strict=False):
                    if value is not None:
                        # Format numbers cleanly (remove trailing .0 for integers)
                        if isinstance(value, float) and value == int(value):
                            value = int(value)
                        pairs.append(f"{header}: {value}")
                if pairs:
                    section_lines.append("，".join(pairs) + "。")
                    total_rows += 1

            if len(section_lines) > 1:
                sheet_texts.append("\n".join(section_lines))

        if not sheet_texts:
            raise ExcelParserError(f"Excel file has no readable data: {path.name}")

        extracted_text = "\n\n".join(sheet_texts)
        return ParseResult(
            extracted_text=extracted_text,
            content_hash=sha256(raw_bytes).hexdigest(),
            mime_type=self.mime_type,
            parser_name=self.parser_name,
            metadata={
                "parser_id": "parser.excel",
                "status": "success",
                "extension": path.suffix.lower(),
                "sheet_count": len(wb.sheetnames),
                "row_count": total_rows,
                "char_count": len(extracted_text),
            },
        )
